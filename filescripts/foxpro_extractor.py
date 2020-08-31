import pandas as pd
import numpy as np
import subprocess
import sys
import os
import re
import datetime as dt
import csv
import xlwt
import openpyxl
import xlrd
import tkinter as tk
import tkinter.ttk as ttk
from tkcalendar import Calendar, DateEntry
from difflib import SequenceMatcher
from xlutils.copy import copy
from xlwt import Workbook
from os.path import join, dirname
from dotenv import load_dotenv

def getDate(date1, date2):
    dates = [date1.get_date(), date2.get_date()]
    return(dates)

def statusCallBack(label):
    if os.path.isfile('./output.xls'):
        label.configure(text="Sukses, file yang dibuat berjudul \"output.xls\"")
    else:
        label.configure(text="Gagal")

def UI():
    master = tk.Tk()
    master.title('Excel Generator GL P2M')
    master.geometry('400x200')
    topframe = tk.Frame(master)
    bottomframe = tk.Frame(master)
    
    l1 = tk.Label(topframe, text='Pilih Tanggal Awal: ', font=("Arial", 15))
    l1.grid(row = 0, column= 0, padx=10, pady=5)
    
    cal1 = DateEntry(topframe)
    cal1.grid(row= 0, column= 1, padx=5, pady=5)

    l2 = tk.Label(topframe, text='Pilih Tanggal Akhir:', font=("Arial", 15))
    l2.grid(row= 1, column= 0, padx=10, pady=5)
    
    cal2 = DateEntry(topframe)
    cal2.grid(row= 1, column= 1, padx=5, pady=5)

    l3 = tk.Label(bottomframe, text='')
    l3.grid(row= 3, column=0)

    b = ttk.Button(bottomframe, text='Buat Excel',command=lambda : [getExcel(getDate(cal1, cal2)),statusCallBack(l3)])
    b.grid(row=2, column=0, padx=10)
    topframe.grid(row= 0, column= 0,padx=50, pady=10)
    bottomframe.grid(row= 1, column= 0,padx=50, pady=10)
    master.mainloop()

def getExcel(filters):
    '''
    Pengambilan data untuk ditulis ke bentuk excel(xls) dan csv
    '''
    data = fix_anomaly(getData())
    if checkMapping() == False:
        mapping(data)
    filtered_data = date_filter(filters, data)
    writeExcel(filtered_data)

def date_filter(filter,data):
    lower_date = filter[0]
    upper_date = filter[1]
    filtered_data = []
    for i in data:
        if i[1] >= lower_date and i[1] <= upper_date:
            filtered_data.append(i)
    return filtered_data

def checkMapping():
    status = False
    try:
        entries = getMapping(raw=True)
        status = True if len(entries) >= 1 else False
    except FileNotFoundError:
        status = False    
    return status

def getMapping(raw=True):
    entries = []
    with open('maptable.csv', mode='r') as maptable:
        reader = csv.reader(maptable, delimiter=',')   
        for row in reader:
            entries.append(row)

    map_table = entries
    mapping = {}
    for i in range(len(map_table)):
        datum = map_table[i]
        for j in range(len(datum)):
            if j == 0:
                continue
            else:
                mapping[datum[j]] = datum[0]

    return entries if raw else mapping

def writeExcel(data):
    '''
    Penulisan entry ke bentuk excel (xls) dan csv
    dengan format biasa untuk csv dan format sesuai journal untuk xls
    '''
    mapping = getMapping(raw=False)
    keys = list(mapping.keys())
    wb = xlwt.Workbook()
    sh = wb.add_sheet('sheet1')
    
    for i in range(len(data)):
        ui_acc = "'"+mapping[str(data[i][2])] if str(data[i][2]) in keys else ''
        entry1 = ['','Universitas Indonesia','','UKK-FT-UP2M','IDR',str(data[i][1]),"'10407", "'00000000","'71",'',ui_acc,'',"'000","'000"
            ,data[i][4],data[i][5],str(data[i][1].strftime('%b'))+'-'+str(data[i][1].year)[2:], 'UKK-FT UP2M ' + data[i][0], data[i][3],'',data[i][3]
            ,'','','','','','','','J','']
        for j in range(len(entry1)):
            sh.write(i,j,entry1[j])
    wb.save('output.xls')

def mapping(data):
    wb = openpyxl.load_workbook(filename='table.xlsx')
    s = wb.active
    m_row = s.max_row
    gl_acc_list = []
    for i in range(2, m_row+1):
        cell = s.cell(row=i, column=2)
        gl_acc_list.append(cell.value)
    wb.close
    
    wb = openpyxl.load_workbook(filename='april.xlsx')
    s = wb.active
    m_row = s.max_row
    ui_acc_apr = []
    for i in range(1, m_row+1):
        ui_acc_apr.append([s.cell(i,5).value, s.cell(i, 10).value, s.cell(i, 18).value
        , s.cell(i, 14).value if s.cell(i, 14).value != None else 0, s.cell(i, 15).value if s.cell(i, 15).value != None else 0])
    wb.close
    
    days = set()     
    for i in ui_acc_apr:
        days.add(i[0].day)
    
    gl_acc_apr = []
    for i in data:
        if i[1].month == 4 and i[1].year == 2020 and i[1].day in days:
            gl_acc_apr.append(i)

    ui_to_gl = {}
    for ui in ui_acc_apr:
        for gl in gl_acc_apr:
            term1 = int(float(gl[4])) == ui[3]
            term2 = int(float(gl[5])) == ui[4]
            desc_gl = gl[3].lower() if gl[3] != None else ''
            desc_ui = ui[2].lower() if ui[2] != None else ''
            term3 = similar(desc_gl, desc_ui) >= 0.75
            if term1 and term2: # Add term 3 if necessary
                if ui[1] not in ui_to_gl.keys():
                    ui_to_gl[ui[1]] = [gl[2]]
                else:
                    ui_to_gl[ui[1]].append(gl[2])
    keys = list(ui_to_gl.keys())
    table_map = []
    for i in range(len(keys)):
        val = ui_to_gl[keys[i]]
        ui_to_gl[keys[i]] = list(set(val))
        # print('Key: {}, Value: {}'.format(keys[i], str(ui_to_gl[keys[i]])))
        entry = ui_to_gl[keys[i]] if type(ui_to_gl[keys[i]]) is list else [ui_to_gl[keys[i]]]
        # print(entry)
        entry.insert(0, keys[i])
        table_map.append(entry)

    with open('maptable.csv', mode='w', newline='') as maptable:
        writer = csv.writer(maptable)
        for i in table_map:
            writer.writerow(i)

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def getData():
    '''
    Mengambil data dari parameter yang didefinisikan .env
    dengan slicing string sesuai dengan divisi jurnal (sekaligus penulisan salah yang ada),
    mengembalikan list dengan isi [data], divisi
    '''
    dotenv_path = join(dirname(__file__), '.env')
    load_dotenv(dotenv_path)

    CST_NAME = os.environ.get('CST_NAME')

    #Change DTJUR.cst to DTJUR.csv (HARUS HAPUS 2 ROW ATAS DULU SECARA MANUAL)
    if os.path.isfile('./'+CST_NAME+'.CST'):
        filename = CST_NAME+'.CST'
        base = os.path.splitext(filename)[0]
        os.rename(filename, base + '.csv')

    new_filename = CST_NAME + '.csv'
    data = pd.read_csv(new_filename, header=None)
    
    string_dump = ''
    for i in range(data.shape[0]):
        string_dump += data.loc[i].values[0]
    splitted_dump = re.split('(SEKRET|SEKRE|SKRET|PP-WEL|P-WEL|PTHN|PLTHN|PP-OTO|PP-BM)', string_dump)
    
    records = {'SEKRET':[], 'PP-WEL':[], 'PLTHN':[], 'PP-OTO':[], 'PP-BM':[]}
    for i in range(len(splitted_dump)):
        if splitted_dump[i] == 'SEKRE':
            splitted_dump[i] = 'SEKRET'
        if splitted_dump[i] == 'SKRET':
            splitted_dump[i] = 'SEKRET'
        if splitted_dump[i] == 'P-WEL':
            splitted_dump[i] = 'PP-WEL'
        if splitted_dump[i] == 'PTHN':
            splitted_dump[i] = 'PLTHN'
            
        if splitted_dump[i] in records.keys():
            records[splitted_dump[i]].append(splitted_dump[i-1])

    entries = []
    for key in records.keys():
        if '' in records[key]:
            records[key].remove('')
        for el in records[key]:
            entry = re.split('\s{2,}', el)
            if '' in entry:
                entry.remove('')
            entry.append(key)
            entries.append(entry)
    return entries

def fix_anomaly(data):

    '''
    Ditemukan banyak anomali data dari fungsi getData()
    data normal memiliki panjang 6, data anomali memiliki panjang tidak 6
    fungsi ini digunakan sebagai "pembersih" untuk data tersebut
    fungsi ini memisahkan dan memfilter data sesuai dengan panjangnya
    Setelah itu untuk setiap data pada posisi ke 2 terdapat atribut tanggal yang tergabung dengan atribut lain
    atribut tersebut dipisah dan diletakkan di posisi ke 2 dan ke 3 
    '''
    cnt = 0
    cnt1 = 0
    # Check anomaly data
    anomalies_less = []
    anomalies_more = []
    for i in range(len(data)):    
        # Pengecekan data > 6
        if len(data[i]) > 6:
            if len(data[i]) <= 7:
                anomalies_more.append(i)
            cnt += 1
        # Pengecekan data < 6
        elif len(data[i]) < 6:
            anomalies_less.append(i)
            cnt1 += 1
    # Fix anomaly data
    for i in range(len(data)):
        if len(data[i]) == 7:
            tmp = data[i].pop(2) + ' ' + data[i].pop(2)
            data[i].insert(2, tmp)
        if len(data[i]) < 6:
            tmp = data[i].pop(0).split()
            for j in tmp[::-1]:
                data[i].insert(0, j)
            if len(data[i]) > 6 or (len(data[i]) == 6 and data[i][0] == 'AC'):
                tmp = data[i].pop(0) + '-' + data[i].pop(0)
                data[i].insert(0, tmp)
            if len(data[i]) < 6:
                tmp = re.split('(KK|KM)', data[i].pop(0))
                if len(tmp) > 2:
                    tmp2 = [tmp[0] + tmp[1], tmp[2]]
                for j in tmp2[::-1]:
                    data[i].insert(0, j)
    # re-check
    anomalies_unique = []
    for i in range(len(data)):
        if len(data[i]) != 6:
            anomalies_unique.append(i)
    # split date
    error_date = []
    for i in range(len(data)):
        if i in anomalies_unique:
            continue
        else:
            tmp = data[i].pop(1)
            try:
                date = dt.datetime.strptime(tmp[0:8], '%Y%m%d').date()
            except ValueError:
                error_date.append(i)
                data[i].insert(1, tmp)
                continue
            account = tmp[8:]
            data[i].insert(1, account)
            data[i].insert(1, date)
    # case khusus karena terdapat data yang salah tanggal
    for i in range(len(error_date)):
        tmp = ''
        if i == 0:
            data[error_date[i]][1] = '20141216120.103'
            tmp = data[error_date[i]].pop(1)
        elif i == 1: 
            data[error_date[i]][1] = '20171214110.101'
            tmp = data[error_date[i]].pop(1)
        elif i == 2:
            data[error_date[i]][1] = '20190109110.101'
            tmp = data[error_date[i]].pop(1)
        elif i == 4:
            data[error_date[i]][1] = '20160423400.105'
            tmp = data[error_date[i]].pop(1)
        else:
            tmp = data[error_date[i]].pop(0).split()
            data[error_date[i]].insert(0, tmp.pop(0))
            tmp = tmp[0]
        date = dt.datetime.strptime(tmp[0:8], '%Y%m%d').date()
        account = tmp[8:]
        data[error_date[i]].insert(1, account)
        data[error_date[i]].insert(1, date)
    # Case khusus untuk beberapa data yang tidak ter split normal
    for i in range(len(anomalies_unique)):
        if i == 0:
            broken_data = data.pop(anomalies_unique[i])
            tmp = []
            tmp_wrapper = []
            cnt = 0
            for j in range(len(broken_data)):
                tmp.append(broken_data[j])
                cnt += 1
                if cnt % 5 == 0:
                    tmp.append('SEKRET')
                    tmp_wrapper.append(tmp)
                    tmp = []
            for k in tmp_wrapper:
                split_this = k.pop(1)
                date = dt.datetime.strptime(split_this[0:8], "%Y%m%d").date()
                acc = split_this[8:]
                k.insert(1, acc)
                k.insert(1, date)
                data.insert(anomalies_unique[i], k)
    # Pembenaran tanggal untuk case khusus diatas
    cnt = 0
    for i in data:
        if len(i) > 7:
            if cnt < 4:
                i.pop(0)
                i.pop(0)
                tmp = i.pop(1)
                date = dt.datetime.strptime(tmp[0:8], "%Y%m%d").date()
                acc = tmp[8:]
                i.insert(1, acc)
                i.insert(1, date)
            elif cnt == 4:
                i.pop(0)
                i.pop(0)
                i.pop(0)
                i.pop(0)
                tmp = i.pop(1)
                date = dt.datetime.strptime(tmp[0:8], "%Y%m%d").date()
                acc = tmp[8:]
                i.insert(1, acc)
                i.insert(1, date)
            else:
                tmp = i.pop(3) + ' ' + i.pop(3)
                i.insert(3, tmp)
            cnt += 1

    #fixing busted account numbers 
    busted_acc = ['00.102', '00.125', '00.137', '10.107','500107','99.999','110101', '110107']
    
    for i in range(len(data)):
        acc = data[i].pop(2)
        if acc == busted_acc[0]:
            data[i].insert(2, '300.102')
        elif acc == busted_acc[1]:
            data[i].insert(2, '500.125')
        elif acc == busted_acc[2]:
            data[i].insert(2, '500.137')
        elif acc == busted_acc[3]:
            data[i].insert(2, '110.107')
        elif acc == busted_acc[4]:
            data[i].insert(2, '500.107')
        elif acc == busted_acc[5]:
            data[i].insert(2, '999.999')
        elif acc == busted_acc[6]:
            data[i].insert(2, '110.101')
        elif acc == busted_acc[7]:
            data[i].insert(2, '110.107')
        else:
            data[i].insert(2, acc)

    return data

if __name__ == "__main__":
    UI()